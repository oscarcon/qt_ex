import os
import glob
from datetime import datetime as dt
from openpyxl import Workbook, load_workbook
from pprint import pprint

class FileHandler:
    def __init__(self):
        pass
    def file_filter(self, folder_path, ext='*', time=''):
        d = {}
        for path in glob.glob(os.path.join(folder_path, f'**/*.{ext.replace(".", "")}'), recursive=True):
            print(path)
            folderpath, filename = path.rsplit('\\', 1)
            match = True
            if time != '':
                if dt.fromtimestamp(os.path.getmtime(path)).strftime('%d/%m/%Y') != time:
                    match = False
            if match:
                if filename not in d.keys():
                    d[filename] = []
                d[filename].append(folderpath)
        return d
    def export_to_excel(self, data, excel_path='output.xlsx'):
        wb = Workbook()
        ws = wb.active
        i = 1
        for key, values in data.items():
            for value in values:
                ws[f'A{i}'] = key
                ws[f'B{i}'] = value
                i += 1
        wb.save(filename=excel_path)
        wb.close()
    def import_from_excel(self, excel_path='output.xlsx'):
        wb = load_workbook(filename=excel_path)
        ws = wb.active
        i = 1
        d = {}
        while True:
            key = ws[f'A{i}'].value
            value = ws[f'B{i}'].value
            if key == None:
                break
            if key not in d:
                d[key] = []
            d[key].append(value)
            i += 1
        return d
            
if __name__ == '__main__':
    file_handler = FileHandler()
    d = file_handler.file_filter('D:\\HiCS', ext='cpp', time='14/05/2020')
    pprint(d)
    file_handler.export_to_excel(d)
    d = file_handler.import_from_excel()
    pprint(d)
